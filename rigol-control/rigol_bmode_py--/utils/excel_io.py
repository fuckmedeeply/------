from __future__ import annotations
from typing import List, Tuple, Union

# Excel saving: requires openpyxl. If not installed, raise a clear message.
def write_wave_excel(filepath: str, metadata_pairs: List[Tuple[str, Union[str, float, int]]], time_s, voltage_v,
                    *, sheet_meta: str="元数据", sheet_wave: str="波形数据"):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise ModuleNotFoundError("openpyxl 未安装：请先执行 `python -m pip install openpyxl`") from e

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = sheet_meta
    ws_meta.append(["参数", "值"])
    for k, v in metadata_pairs:
        ws_meta.append([k, v])

    ws_wave = wb.create_sheet(sheet_wave)
    ws_wave.append(["时间(s)", "电压(V)"])
    for t, y in zip(time_s, voltage_v):
        ws_wave.append([float(t), float(y)])

    for ws in (ws_meta, ws_wave):
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    wb.save(filepath)
